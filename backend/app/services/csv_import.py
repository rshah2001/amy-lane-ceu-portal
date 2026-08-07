import csv
import io
import xml.etree.ElementTree as ElementTree
import zipfile
from zipfile import BadZipFile
from datetime import datetime
from decimal import Decimal, DivisionByZero, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from PIL import Image, ImageOps
from pypdf import PdfReader
import pytesseract
from sqlalchemy import delete, select, update
from sqlalchemy.orm import Session

from app.models.event_attendee import EventAttendee
from app.services import storage
from app.models.survey_result import SurveyResult
from app.models.test_result import TestResult
from app.services.attendee_match import EventRoster, get_or_create_link, match_or_create_attendee
from app.services.compliance import recalculate_event
from app.services.identity import (
    core_name,
    humanize_name,
    names_are_variants,
    normalize_email,
    split_name,
)

ALIASES = {
    "full_name": ["full name", "name", "attendee", "participant name", "student name"],
    "first_name": ["first name", "firstname", "given name", "registration first name"],
    "last_name": ["last name", "lastname", "surname", "family name", "registration last name"],
    "email": ["email", "email address", "e-mail", "participant email", "registration email"],
    "company": ["company", "organization", "dealer", "employer"],
    "license_number": ["license number", "license", "credential id"],
    "score": ["score", "test score", "percentage", "percent", "grade"],
    "completed": ["completed", "complete", "submitted", "status"],
    "completed_at": ["completed at", "completion date", "submitted at", "timestamp"],
}

# Sheet-format hints the uploader can optionally send with the file. "other"
# (or no hint at all) keeps the default extension-based parsing untouched.
SHEET_FORMAT_SUFFIXES = {
    "spreadsheet": {".csv", ".xlsx"},
    "word": {".docx"},
    "virtual_meeting": {".csv", ".xlsx"},
}
SHEET_FORMAT_LABELS = {
    "spreadsheet": "Spreadsheet (Excel/CSV)",
    "word": "Word document",
    "virtual_meeting": "Virtual meeting export (Zoom/Teams)",
}
SHEET_FORMATS = set(SHEET_FORMAT_SUFFIXES) | {"other"}

NO_NAMES_MESSAGE = (
    "No attendee names could be read from this file — check the sheet format "
    'selection or convert it to a spreadsheet (CSV/XLSX) with a "Name" column.'
)
# Sibling of NO_NAMES_MESSAGE for the case where names *were* read but every
# row was rejected. Both mean the same thing to the caller: nothing was
# imported, so nothing on the event was touched.
NOTHING_IMPORTED_MESSAGE = (
    "No rows from this file could be imported, so the results already recorded "
    "for this event were left unchanged. Fix the row problems listed below and "
    "upload the file again."
)

# Score-basis hints the uploader can optionally send with a post-test file,
# mirroring the sheet_format picker. An explicit choice always beats inference.
SCORE_BASIS_LABELS = {
    "percent": "percentages (0-100)",
    "out_of_10": "scores out of 10 (8 = 80%)",
    "out_of_1": "fractions of 1 (0.85 = 85%)",
}
SCORE_BASES = set(SCORE_BASIS_LABELS)

AMBIGUOUS_SCORE_MESSAGE = (
    "The score column is ambiguous: every value is between 0 and 10 and none "
    'says "%" or "x/10", so 8 could mean 8% (a fail) or 8/10 (a pass) and '
    "guessing would issue certificates for failed tests. Re-upload the file "
    'with an explicit "%" or "x/10" on every row, or choose the score basis '
    "(percent, out_of_10 or out_of_1) on the upload."
)


def sample_scores(values: list[str], limit: int = 6) -> list[str]:
    """A few distinct unit-less values from the score column, in file order."""
    seen: list[str] = []
    for value in values:
        text = (value or "").strip()
        if not text or _bare_number(text) is None or text in seen:
            continue
        seen.append(text)
        if len(seen) == limit:
            break
    return seen


class AmbiguousScoreError(ValueError):
    """The score column has no unit and could mean two different things.

    Carries the offending values so the caller can show the uploader what is
    actually in their file. Subclasses ValueError because every existing caller
    already handles that — this only adds detail, it does not change control
    flow.
    """

    def __init__(self, samples: list[str]) -> None:
        self.samples = samples
        detail = f" Values found: {', '.join(samples)}." if samples else ""
        super().__init__(AMBIGUOUS_SCORE_MESSAGE + detail)


class EmptyImportError(ValueError):
    """Raised when an import would replace existing results with nothing.

    The destructive reset that precedes an import (clearing scores, deleting
    file-sourced results) must never run for a file that turns out to import
    zero rows — that is how a mis-typed re-upload wiped a whole event's scores.
    Raising instead of returning keeps the caller from recording a successful
    upload; the rows that failed travel along in ``errors``.
    """

    def __init__(self, message: str, row_count: int, errors: list[dict[str, Any]]) -> None:
        super().__init__(message)
        self.row_count = row_count
        self.errors = errors


def _canonical_header(value: str) -> str:
    return " ".join(value.strip().lower().replace("_", " ").replace("-", " ").split())


def _normalize_row(row: dict[str, str]) -> dict[str, str]:
    """Canonical-header view of a row, built once and reused for every lookup."""
    return {
        _canonical_header(key): ("" if value is None else str(value).strip())
        for key, value in row.items()
        if key
    }


def _lookup(values: dict[str, str], field: str) -> str | None:
    for alias in ALIASES[field]:
        if values.get(alias):
            return values[alias]
    return None


def _has_column(values: dict[str, str], field: str) -> bool:
    """Whether the file has a column for this field at all (value may be blank).

    Column *presence* and cell *emptiness* mean different things: a survey
    export with no "Completed" column is itself the list of people who
    completed it, while a blank cell in a Completed column means they did not.
    """
    return any(alias in values for alias in ALIASES[field])


def _identity(values: dict[str, str]) -> dict[str, str | None]:
    first = _lookup(values, "first_name")
    last = _lookup(values, "last_name")
    full = _lookup(values, "full_name") or " ".join(value for value in [first, last] if value)
    if not full:
        raise ValueError('A name is required (expected a "Name" or "First/Last Name" column)')
    full = humanize_name(full)
    split_first, split_last = split_name(full)
    return {
        "full_name": full,
        "first_name": first or split_first,
        "last_name": last or split_last,
        "email": _lookup(values, "email"),
        "company": _lookup(values, "company"),
        "license_number": _lookup(values, "license_number"),
    }


def _bare_number(value: str) -> Decimal | None:
    """The numeric value of a cell that states no unit, else None.

    A cell carrying "%" or "x/10" says what it means and is parsed on its own
    terms; only these unit-less cells need a column-wide interpretation.
    """
    if "%" in value or "/" in value:
        return None
    try:
        return Decimal(value.strip())
    except InvalidOperation:
        return None


def resolve_score_basis(chosen: str | None, values: list[str]) -> str | None:
    """Decide ONCE, for the whole column, what a unit-less score means.

    Deciding per cell is how an 8 meaning 8% became 80.0 — exactly the pass
    mark — and printed a certificate for a failed test. So:

    - an explicit choice from the uploader always wins;
    - a "%" anywhere in the column, or any value above 10, makes the whole
      column percentages;
    - a column whose values all sit between 0 and 10 with no "%" and no
      fraction is genuinely ambiguous and is REFUSED, never guessed;
    - a column with no unit-less values at all needs no basis (None): every
      cell already says what it means.
    """
    if chosen:
        if chosen not in SCORE_BASES:
            raise ValueError(f"Unknown score basis: {chosen!r}")
        return chosen
    numbers = [number for number in (_bare_number(value) for value in values) if number is not None]
    if not numbers:
        return None
    if any("%" in value for value in values) or any(number > 10 for number in numbers):
        return "percent"
    # Name the values that caused the refusal. The uploader is looking at a
    # dialog, not at the spreadsheet, and "8, 9, 7.5" is what lets them answer
    # "out of 10" without going back to Excel to check what is in the column.
    raise AmbiguousScoreError(sample_scores(values))


def describe_score_basis(basis: str | None, chosen: bool) -> str:
    source = "as selected on this upload" if chosen else "read from the file"
    if basis is None:
        return (
            "Interpreted scores exactly as written — every row stated its own "
            'unit ("%" or "x/10").'
        )
    return (
        f"Interpreted scores as {SCORE_BASIS_LABELS[basis]}, {source}. "
        "Re-upload with the correct score basis if that is not what the file meant."
    )


def _parse_score(value: str | None, basis: str | None = None) -> Decimal:
    """Normalize one score cell to a 0-100 percentage.

    Accepted: an explicit percentage ("85%"), an explicit fraction ("8/10",
    "17/20"), or a unit-less number read on the column's ``basis`` (see
    resolve_score_basis). A unit-less number is never interpreted on its own.
    """
    if value is None:
        raise ValueError(
            'A post-test score is required (expected a "Score" column with a value like 85%, 8/10, or 8)'
        )
    is_percent = "%" in value
    text = value.replace("%", "").strip()
    try:
        if "/" in text:
            numerator, _, denominator = text.partition("/")
            score = Decimal(numerator.strip()) / Decimal(denominator.strip()) * 100
        else:
            score = Decimal(text)
            if not is_percent:
                if basis is None:
                    raise ValueError(AMBIGUOUS_SCORE_MESSAGE)
                if basis == "out_of_10":
                    score *= 10
                elif basis == "out_of_1":
                    score *= 100
    except (InvalidOperation, DivisionByZero) as exc:
        raise ValueError(
            f"Invalid score: {value!r} (expected a percentage like 85%, a fraction like 8/10, or a score out of 10)"
        ) from exc
    if score < 0 or score > 100:
        raise ValueError(f"Post-test score must be between 0 and 100, got {score}")
    return score.quantize(Decimal("0.01"))


# Words that state an attendee did NOT take part. "n/a" counts: on a survey
# sheet it is how presenters write "no response", not "yes".
NOT_COMPLETED_VALUES = {
    "false",
    "no",
    "n",
    "0",
    "incomplete",
    "not completed",
    "did not attend",
    "did not complete",
    "dna",
    "absent",
    "no show",
    "no-show",
    "n/a",
    "na",
    "none",
    "-",
}


def _parse_completed(values: dict[str, str]) -> bool:
    """Read the completion flag for one row.

    A file with no completion column at all IS the list of completions (that
    is how survey exports arrive), so a missing column means completed. A
    column that exists but is blank on this row does not: a human left it
    empty, which is never a statement that someone completed anything.
    """
    if not _has_column(values, "completed"):
        return True
    value = _lookup(values, "completed")
    if value is None:
        return False
    return value.strip().lower() not in NOT_COMPLETED_VALUES


def _parse_datetime(value: str | None) -> datetime | None:
    if not value:
        return None
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        return None


def _decode_csv(contents: bytes) -> str:
    """Decode CSV bytes, handling UTF-8/UTF-16 BOMs and legacy Excel encodings."""
    if contents.startswith((b"\xff\xfe", b"\xfe\xff")):
        return contents.decode("utf-16")
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            return contents.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("The file is not text in a supported encoding (UTF-8 or Windows-1252)")


def looks_like_csv_text(contents: bytes) -> bool:
    """Whether the upload gate should let these bytes through as a CSV.

    Kept next to _decode_csv on purpose: the gate must accept exactly what the
    decoder can read. Judging a CSV by "does it decode as UTF-8" rejected the
    cp1252 and UTF-16 exports Excel produces — files this importer reads fine.

    The gate still has to keep real binaries out, and latin-1 "decodes"
    anything, so the decoded text is checked for the control characters that
    text never contains but compressed/binary payloads always do.
    """
    if not contents:
        return False
    try:
        text = _decode_csv(contents)
    except ValueError:
        return False
    # Tab, CR and LF are legitimate in CSV; every other C0/C1 control byte (and
    # DEL) means this is not a spreadsheet export.
    sample = text[:8192]
    return not any(
        (code < 32 and char not in "\t\r\n") or 127 <= code < 160
        for char, code in ((char, ord(char)) for char in sample)
    )


def _dedupe_headers(fieldnames: list[str]) -> list[str]:
    """Rename repeated header names so duplicate columns cannot clobber each other.

    csv.DictReader keeps only the last column for a repeated header; renaming
    later duplicates ("Email" -> "Email (2)") preserves the first column, which
    is the one the alias lookup should use.
    """
    seen: dict[str, int] = {}
    unique: list[str] = []
    for name in fieldnames:
        name = (name or "").strip()
        canonical = _canonical_header(name)
        count = seen.get(canonical, 0)
        seen[canonical] = count + 1
        unique.append(f"{name} ({count + 1})" if name and count else name)
    return unique


def parse_csv_bytes(contents: bytes) -> list[dict[str, str]]:
    text = _decode_csv(contents)
    reader = csv.DictReader(io.StringIO(text))
    if reader.fieldnames is None:
        raise ValueError("The file is empty")
    reader.fieldnames = _dedupe_headers(list(reader.fieldnames))
    return list(reader)


def _detect_header(rows: list[tuple], prefer_virtual: bool = False) -> int | None:
    """Find the header row: the first row with >=2 labels that names a person or email.

    Skips leading title/summary rows (e.g. Microsoft Teams "1. Summary" blocks).
    With `prefer_virtual` (the uploader said this is a Zoom/Teams export), first
    look for the attendee table proper — a row naming both a person column and
    an email/join-time column — so summary lines like "Meeting name" can't win.
    """
    if prefer_virtual:
        for index, row in enumerate(rows):
            cells = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
            if len(cells) < 2:
                continue
            canonical = [_canonical_header(cell) for cell in cells]
            if any("name" in cell for cell in canonical) and any(
                "email" in cell or "join time" in cell for cell in canonical
            ):
                return index
    for index, row in enumerate(rows):
        cells = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
        if len(cells) < 2:
            continue
        canonical = [_canonical_header(cell) for cell in cells]
        if any("email" in cell or "name" in cell for cell in canonical):
            return index
    return None


def _records_from_rows(rows: list[tuple], header_index: int) -> list[dict[str, str]]:
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[header_index]]
    records: list[dict[str, str]] = []
    for row in rows[header_index + 1 :]:
        record = {
            headers[index]: "" if cell is None else str(cell).strip()
            for index, cell in enumerate(row)
            if index < len(headers) and headers[index]
        }
        if any(value for value in record.values()):
            records.append(record)
    return records


def _xlsx_cell_value(cell) -> Any:
    """Excel stores a percent-formatted 85% as 0.85; surface it as "85%" so
    the score parser doesn't misread it as a fraction of one percent."""
    value = cell.value
    if (
        isinstance(value, (int, float))
        and not isinstance(value, bool)
        and "%" in (cell.number_format or "")
    ):
        return f"{value * 100:g}%"
    return value


def parse_xlsx_bytes(contents: bytes, prefer_virtual: bool = False) -> list[dict[str, str]]:
    workbook = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    # Real exports (e.g. Microsoft Teams) ship multiple sheets: a summary sheet
    # plus the attendee table. Choose the sheet whose header best identifies people.
    best: tuple[tuple[int, int, int], list[dict[str, str]]] | None = None
    for worksheet in workbook.worksheets:
        rows = [tuple(_xlsx_cell_value(cell) for cell in row) for row in worksheet.iter_rows()]
        header_index = _detect_header(rows, prefer_virtual=prefer_virtual)
        if header_index is None:
            continue
        records = _records_from_rows(rows, header_index)
        if not records:
            continue
        canonical = [_canonical_header(str(cell)) for cell in rows[header_index] if cell]
        score = (
            sum("email" in cell for cell in canonical),
            sum("name" in cell for cell in canonical),
            len(records),
        )
        if best is None or score > best[0]:
            best = (score, records)
    if best is None:
        raise ValueError(
            "no attendee table found (expected a sheet with a header row naming a name or email column)"
        )
    return best[1]


def _split_ocr_line(line: str) -> list[str]:
    if "," in line:
        return next(csv.reader([line]))
    if "|" in line:
        return [part.strip() for part in line.split("|")]
    if "\t" in line:
        return [part.strip() for part in line.split("\t")]
    # Best-effort fallback for printed table scans. Commas/pipes produce better OCR.
    import re

    return [part.strip() for part in re.split(r"\s{2,}", line) if part.strip()]


def _parse_ocr_text(text: str) -> tuple[list[dict[str, str]], list[dict[str, Any]]]:
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    if not lines:
        return [], [{"row": 0, "message": "OCR found no readable text"}]

    header_index = next(
        (
            index
            for index, line in enumerate(lines)
            if any(alias in _canonical_header(line) for alias in ["email", "name", "score"])
        ),
        0,
    )
    headers = _split_ocr_line(lines[header_index])
    if len(headers) < 2:
        return [], [
            {
                "row": 0,
                "message": "OCR text was found, but a table header could not be detected. Use CSV/XLSX or a clearer comma-separated image.",
            }
        ]

    records: list[dict[str, str]] = []
    errors: list[dict[str, Any]] = [
        {
            "row": 0,
            "message": "Image OCR is best-effort. Review extracted rows before approving certificates.",
        }
    ]
    for row_number, line in enumerate(lines[header_index + 1 :], start=2):
        parts = _split_ocr_line(line)
        if len(parts) < len(headers):
            errors.append({"row": row_number, "message": f"OCR skipped low-confidence row: {line}"})
            continue
        records.append({header: parts[index] for index, header in enumerate(headers) if header})
    return records, errors


def _ocr_text_from_image(contents: bytes) -> str:
    """OCR raw image bytes into text. Shared by image uploads and scanned PDFs."""
    try:
        image = Image.open(io.BytesIO(contents))
        image = ImageOps.autocontrast(ImageOps.grayscale(image))
        if image.width < 1400:
            scale = 1400 / image.width
            image = image.resize(
                (1400, max(1, round(image.height * scale))),
                Image.Resampling.LANCZOS,
            )
        return pytesseract.image_to_string(image, config="--psm 6")
    except (OSError, pytesseract.TesseractError) as exc:
        raise ValueError(f"Image OCR failed: {exc}") from exc


def parse_image_bytes(contents: bytes) -> tuple[list[dict[str, str]], list[dict[str, Any]]]:
    return _parse_ocr_text(_ocr_text_from_image(contents))


def parse_pdf_bytes(contents: bytes) -> tuple[list[dict[str, str]], list[dict[str, Any]]]:
    """Extract attendee rows from a PDF sign-in sheet.

    Pages with a text layer go through the shared line parser; pages that are
    embedded scans (scan-to-PDF, the common case for sign-in sheets) fall back
    to OCR on each embedded image.
    """
    try:
        reader = PdfReader(io.BytesIO(contents))
        if reader.is_encrypted:
            reader.decrypt("")
        pages = list(reader.pages)
    except Exception as exc:  # pypdf raises many exception types on bad input
        raise ValueError(f"Invalid PDF file: {exc}") from exc

    records: list[dict[str, str]] = []
    warnings: list[dict[str, Any]] = [
        {
            "row": 0,
            "message": (
                "PDF text/OCR extraction is best-effort; a CSV or XLSX export is more "
                "reliable. Review extracted rows before approving certificates."
            ),
        }
    ]
    seen_messages = {warnings[0]["message"]}

    def _merge(new_records: list[dict[str, str]], new_warnings: list[dict[str, Any]]) -> None:
        records.extend(new_records)
        for warning in new_warnings:
            if warning.get("message") not in seen_messages:
                seen_messages.add(warning.get("message"))
                warnings.append(warning)

    for page_number, page in enumerate(pages, start=1):
        try:
            text = page.extract_text() or ""
        except Exception:
            text = ""
        if text.strip():
            _merge(*_parse_ocr_text(text))
            continue
        # No text layer: OCR any embedded page images (scan-to-PDF case).
        try:
            images = list(page.images)
        except Exception:
            images = []
        if not images:
            _merge([], [{"row": 0, "message": f"PDF page {page_number} has no extractable text or images."}])
            continue
        for image in images:
            try:
                _merge(*_parse_ocr_text(_ocr_text_from_image(image.data)))
            except ValueError as exc:
                _merge([], [{"row": 0, "message": f"PDF page {page_number}: {exc}"}])
    if not records:
        raise ValueError(
            "no attendee rows could be extracted from the PDF (a CSV or XLSX export is more reliable)"
        )
    return records, warnings


_DOCX_W = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"


def _docx_text(element) -> str:
    return "".join(node.text or "" for node in element.iter(f"{_DOCX_W}t")).strip()


def parse_docx_bytes(contents: bytes) -> tuple[list[dict[str, str]], list[dict[str, Any]]]:
    """Extract attendee rows from a Word (.docx) sign-in sheet.

    Most sign-in sheets are tables: read w:tbl/w:tr/w:tc from word/document.xml
    (stdlib zipfile + ElementTree; no python-docx needed) with the first
    header-looking row as column names. Documents without a usable table fall
    back to paragraph text lines through the shared line parser.
    """
    try:
        with zipfile.ZipFile(io.BytesIO(contents)) as archive:
            document = archive.read("word/document.xml")
        root = ElementTree.fromstring(document)
    except (BadZipFile, KeyError, ElementTree.ParseError, OSError) as exc:
        raise ValueError(f"Invalid DOCX file: {exc}") from exc

    best: tuple[tuple[int, int, int], list[dict[str, str]]] | None = None
    for table in root.iter(f"{_DOCX_W}tbl"):
        rows = [
            tuple(_docx_text(cell) for cell in row.findall(f"{_DOCX_W}tc"))
            for row in table.findall(f"{_DOCX_W}tr")
        ]
        if len(rows) < 2:
            continue
        # First row is normally the header; _detect_header also skips any
        # leading title rows, matching the XLSX behavior.
        header_index = _detect_header(rows)
        if header_index is None:
            header_index = 0
        records = _records_from_rows(rows, header_index)
        if not records:
            continue
        canonical = [_canonical_header(str(cell)) for cell in rows[header_index] if cell]
        score = (
            sum("email" in cell for cell in canonical),
            sum("name" in cell for cell in canonical),
            len(records),
        )
        if best is None or score > best[0]:
            best = (score, records)
    if best is not None:
        return best[1], []

    # No table: treat paragraph lines like OCR/plain text.
    lines = [_docx_text(paragraph) for paragraph in root.iter(f"{_DOCX_W}p")]
    records, warnings = _parse_ocr_text("\n".join(line for line in lines if line))
    if not records:
        raise ValueError(
            "no attendee rows found in the DOCX (expected a table with a name or email column)"
        )
    return records, warnings


def _parse_virtual_csv(contents: bytes) -> list[dict[str, str]]:
    """CSV parsing for virtual-meeting attendance exports.

    Teams/Zoom reports often put summary lines above the real header, which the
    default DictReader path would mistake for column names. Locate the actual
    attendee header row first; if none looks virtual, fall back to the default.
    """
    text = _decode_csv(contents)
    rows = [tuple(row) for row in csv.reader(io.StringIO(text))]
    header_index = _detect_header(rows, prefer_virtual=True)
    if header_index is None:
        return parse_csv_bytes(contents)
    return _records_from_rows(rows, header_index)


def parse_document_bytes(
    filename: str,
    contents: bytes,
    sheet_format: str | None = None,
) -> tuple[list[dict[str, str]], list[dict[str, Any]]]:
    suffix = Path(filename).suffix.lower()
    if sheet_format in (None, "", "other"):
        sheet_format = None
    warnings: list[dict[str, Any]] = []
    if sheet_format is not None and suffix not in SHEET_FORMAT_SUFFIXES[sheet_format]:
        # A wrong hint must never sink the upload: warn, then trust the file.
        warnings.append(
            {
                "row": 0,
                "message": (
                    f'The selected sheet format "{SHEET_FORMAT_LABELS[sheet_format]}" does not '
                    f"match the uploaded {suffix} file; the file was read based on its actual type."
                ),
            }
        )
        sheet_format = None
    prefer_virtual = sheet_format == "virtual_meeting"
    if suffix == ".csv":
        try:
            records = _parse_virtual_csv(contents) if prefer_virtual else parse_csv_bytes(contents)
            return records, warnings
        except (csv.Error, UnicodeDecodeError) as exc:
            raise ValueError(f"Invalid CSV file: {exc}") from exc
    if suffix == ".xlsx":
        try:
            return parse_xlsx_bytes(contents, prefer_virtual=prefer_virtual), warnings
        except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
            raise ValueError(f"Invalid XLSX file: {exc}") from exc
    if suffix in {".png", ".jpg", ".jpeg"}:
        records, notes = parse_image_bytes(contents)
        return records, warnings + notes
    if suffix == ".pdf":
        records, notes = parse_pdf_bytes(contents)
        return records, warnings + notes
    if suffix == ".docx":
        records, notes = parse_docx_bytes(contents)
        return records, warnings + notes
    raise ValueError("Supported uploads are CSV, XLSX, PDF, DOCX, PNG, JPG, and JPEG")


class _PreparedRow:
    """One spreadsheet row, parsed and validated before anything is written."""

    __slots__ = ("number", "row", "values", "identity", "norm_email", "score")

    def __init__(
        self,
        number: int,
        row: dict[str, str],
        values: dict[str, str],
        identity: dict[str, str | None],
    ) -> None:
        self.number = number
        self.row = row
        self.values = values
        self.identity = identity
        self.norm_email = normalize_email(identity["email"])
        self.score: Decimal | None = None


def _prepare_rows(
    rows: list[dict[str, str]], errors: list[dict[str, Any]]
) -> list[_PreparedRow]:
    """Read every row's identity before the import touches the database."""
    prepared: list[_PreparedRow] = []
    for row_number, row in enumerate(rows, start=2):
        values = _normalize_row(row)
        # Skip whitespace-only rows (common padding in hand-edited spreadsheets)
        # without disturbing the row numbers reported for later errors.
        if not any(values.values()):
            continue
        try:
            identity = _identity(values)
        except ValueError as exc:
            errors.append({"row": row_number, "message": str(exc)})
            continue
        prepared.append(_PreparedRow(row_number, row, values, identity))
    return prepared


def _reject_indistinguishable_rows(
    prepared: list[_PreparedRow], errors: list[dict[str, Any]]
) -> list[_PreparedRow]:
    """Refuse rows that name the same person twice with no way to tell them apart.

    Two "John Smith" rows with no email on one sign-in sheet used to resolve to
    a single attendee: the roster showed one person, no error was raised, and
    one of the two real people simply never got a certificate. There is no way
    to know whether such rows are two people or one, so neither is imported and
    the uploader is asked for the emails that would settle it.
    """
    groups: dict[str, list[_PreparedRow]] = {}
    for item in prepared:
        # An email is an identity: rows carrying one are never ambiguous.
        if item.norm_email:
            continue
        groups.setdefault(core_name(item.identity["full_name"]), []).append(item)

    rejected: set[int] = set()
    for group in groups.values():
        # A shared core name only *suggests* the same person; confirm with the
        # variant rules so "Bob A. Smith" and "Bob C. Smith" stay two people.
        clusters: list[list[_PreparedRow]] = []
        for item in group:
            for cluster in clusters:
                if names_are_variants(
                    cluster[0].identity["full_name"], item.identity["full_name"]
                ):
                    cluster.append(item)
                    break
            else:
                clusters.append([item])
        for cluster in clusters:
            if len(cluster) < 2:
                continue
            names = sorted({item.identity["full_name"] or "" for item in cluster})
            if len(names) == 1:
                message = (
                    f'{len(cluster)} rows named "{names[0]}" have no email address — '
                    "add emails to tell them apart. None of these rows were imported."
                )
            else:
                spelled = " and ".join(f'"{name}"' for name in names)
                message = (
                    f"Rows named {spelled} have no email address and may be the same "
                    "person — add emails to tell them apart. None of these rows were "
                    "imported."
                )
            for item in cluster:
                rejected.add(item.number)
                errors.append({"row": item.number, "message": message})
    return [item for item in prepared if item.number not in rejected]


def _prepare_scores(
    prepared: list[_PreparedRow],
    score_basis: str | None,
    errors: list[dict[str, Any]],
) -> tuple[list[_PreparedRow], str | None]:
    """Resolve the column's score basis, then read every score with it.

    Scores are parsed here, before the destructive reset, so a file whose
    scores are all unreadable cannot clear the scores already on the event.
    """
    basis = resolve_score_basis(
        score_basis, [_lookup(item.values, "score") or "" for item in prepared]
    )
    kept: list[_PreparedRow] = []
    for item in prepared:
        try:
            item.score = _parse_score(_lookup(item.values, "score"), basis)
        except ValueError as exc:
            errors.append({"row": item.number, "message": str(exc)})
            continue
        kept.append(item)
    return kept, basis


def _reset_previous_results(db: Session, event_id: int, file_type: str) -> None:
    """Clear the results a re-import replaces.

    Destructive by design and therefore only ever called once the import is
    known to have rows to put back (see process_rows).
    """
    # Attendance recorded through the public QR check-in (checked_in_at set)
    # must survive file re-imports: only rows sourced from files are reset.
    if file_type == "registration":
        db.execute(
            update(EventAttendee)
            .where(EventAttendee.event_id == event_id, EventAttendee.checked_in_at.is_(None))
            .values(registered=False)
        )
    elif file_type == "attendance":
        db.execute(
            update(EventAttendee)
            .where(EventAttendee.event_id == event_id, EventAttendee.checked_in_at.is_(None))
            .values(attended=False)
        )
    if file_type == "post_test":
        # Re-importing a file only replaces file-sourced results; submissions
        # made on the public test page ("web") and their flags must survive.
        web_test_attendees = select(TestResult.attendee_id).where(
            TestResult.event_id == event_id, TestResult.source == "web"
        )
        db.execute(
            update(EventAttendee)
            .where(
                EventAttendee.event_id == event_id,
                EventAttendee.attendee_id.not_in(web_test_attendees),
            )
            .values(test_completed=False, test_score=None)
        )
        db.execute(
            delete(TestResult).where(TestResult.event_id == event_id, TestResult.source == "upload")
        )
    elif file_type == "survey":
        # Anonymous web submissions have a NULL attendee_id; a NULL inside the
        # NOT IN subquery would make it match nothing, wiping everyone's flag.
        web_survey_attendees = select(SurveyResult.attendee_id).where(
            SurveyResult.event_id == event_id,
            SurveyResult.source == "web",
            SurveyResult.attendee_id.is_not(None),
        )
        db.execute(
            update(EventAttendee)
            .where(
                EventAttendee.event_id == event_id,
                EventAttendee.attendee_id.not_in(web_survey_attendees),
            )
            .values(survey_completed=False)
        )
        db.execute(
            delete(SurveyResult).where(SurveyResult.event_id == event_id, SurveyResult.source == "upload")
        )


def process_rows(
    db: Session,
    event_id: int,
    file_type: str,
    rows: list[dict[str, str]],
    extraction_errors: list[dict[str, Any]] | None = None,
    score_basis: str | None = None,
) -> tuple[int, list[dict[str, Any]]]:
    """Import parsed rows into an event, replacing what the last import wrote.

    Runs in two phases on purpose. Everything that can reject a row — names,
    duplicate-name ambiguity, the score column's unit — is resolved first,
    against no database state; only then does the destructive reset run, inside
    a SAVEPOINT, so a file that turns out to import nothing leaves the event
    exactly as it was and the caller can fail the request instead of reporting
    a success that quietly erased an evening's results.
    """
    if file_type not in {"registration", "attendance", "post_test", "survey"}:
        raise ValueError(f"Unsupported file type: {file_type}")
    errors: list[dict[str, Any]] = list(extraction_errors or [])

    prepared = _prepare_rows(rows, errors)
    named_rows = len(prepared)
    prepared = _reject_indistinguishable_rows(prepared, errors)
    basis: str | None = None
    if file_type == "post_test" and prepared:
        try:
            prepared, basis = _prepare_scores(prepared, score_basis, errors)
        except ValueError as exc:
            # A score column that cannot be read with certainty stops the
            # import here, before the reset, so the event keeps the scores it
            # already had rather than losing them to a file we refused.
            errors.append({"row": 0, "message": str(exc)})
            raise EmptyImportError(str(exc), len(rows), errors) from exc

    if not prepared:
        # Nothing to import: say so and leave the event untouched. The
        # zero-names wording stays for the case it was written for (a file the
        # parser could read no names from at all).
        message = NOTHING_IMPORTED_MESSAGE if named_rows else NO_NAMES_MESSAGE
        errors.append({"row": 0, "message": message})
        raise EmptyImportError(message, len(rows), errors)

    if file_type == "post_test":
        # Always report how the scores were read: a misread column is the
        # difference between a pass and a fail, and only the admin can catch it.
        errors.append(
            {
                "row": 0,
                "level": "info",
                "message": describe_score_basis(basis, bool(score_basis)),
            }
        )

    # One batched read of everything this file can match, instead of two
    # queries per row against a pooled remote database.
    roster = EventRoster(db, event_id, {item.norm_email for item in prepared if item.norm_email})

    savepoint = db.begin_nested()
    try:
        _reset_previous_results(db, event_id, file_type)
        imported = 0
        for item in prepared:
            identity = item.identity
            row_notes: list[str] = []
            try:
                attendee = match_or_create_attendee(
                    db,
                    event_id,
                    identity["full_name"] or "",
                    identity["email"],
                    first_name=identity["first_name"],
                    last_name=identity["last_name"],
                    company=identity["company"],
                    license_number=identity["license_number"],
                    roster=roster,
                    notes=row_notes,
                )
                link = get_or_create_link(db, event_id, attendee.id, roster=roster)
                if file_type == "registration":
                    link.registered = True
                elif file_type == "attendance":
                    link.attended = True
                elif file_type == "post_test":
                    score = item.score
                    link.test_completed = True
                    link.test_score = score
                    db.add(
                        TestResult(
                            event_id=event_id,
                            attendee_id=attendee.id,
                            score=score,
                            passed=score >= Decimal("80"),
                            completed_at=_parse_datetime(_lookup(item.values, "completed_at")),
                            raw_payload=item.row,
                            source="upload",
                        )
                    )
                elif file_type == "survey":
                    completed = _parse_completed(item.values)
                    link.survey_completed = completed
                    db.add(
                        SurveyResult(
                            event_id=event_id,
                            attendee_id=attendee.id,
                            completed=completed,
                            completed_at=_parse_datetime(_lookup(item.values, "completed_at")),
                            raw_payload=item.row,
                            source="upload",
                        )
                    )
                imported += 1
            except ValueError as exc:
                errors.append({"row": item.number, "message": str(exc)})
                continue
            finally:
                for note in row_notes:
                    errors.append({"row": item.number, "level": "info", "message": note})
        if imported:
            # Links created during the import are still pending (see
            # get_or_create_link); the session does not autoflush, so they must
            # be written before the recalculation queries them back.
            db.flush()
            recalculate_event(db, event_id)
    except Exception:
        savepoint.rollback()
        raise
    if not imported:
        # Belt and braces for the invariant the two phases already enforce: a
        # reset that put nothing back is rolled all the way out.
        savepoint.rollback()
        errors.append({"row": 0, "message": NOTHING_IMPORTED_MESSAGE})
        raise EmptyImportError(NOTHING_IMPORTED_MESSAGE, len(rows), errors)
    savepoint.commit()
    return len(rows), errors


def process_document(
    db: Session,
    event_id: int,
    file_type: str,
    filename: str,
    contents: bytes,
    sheet_format: str | None = None,
    score_basis: str | None = None,
) -> tuple[int, list[dict[str, Any]]]:
    rows, extraction_errors = parse_document_bytes(filename, contents, sheet_format=sheet_format)
    return process_rows(db, event_id, file_type, rows, extraction_errors, score_basis=score_basis)


def process_csv(
    db: Session,
    event_id: int,
    file_type: str,
    contents: bytes,
    score_basis: str | None = None,
) -> tuple[int, list[dict[str, Any]]]:
    return process_rows(
        db, event_id, file_type, parse_csv_bytes(contents), score_basis=score_basis
    )


def save_upload(contents: bytes, destination: Path) -> None:
    # Routed through the storage abstraction: Supabase Storage when configured,
    # plain local disk otherwise. Raises StorageError if a remote write fails —
    # original uploads cannot be regenerated, so the request must fail loudly.
    storage.save_bytes(destination, contents)
